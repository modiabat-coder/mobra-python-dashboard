"""Tests for printable forms, backup packaging, and optional email safeguards."""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from mobra.config import APP_TITLE, AUTHOR_EMAIL, FULL_DISCLAIMER, HOW_TO_USE_STEPS, configured_author_email
from mobra.operational_tools import (
    MAX_EMAIL_ATTACHMENT_BYTES,
    EmailBackupError,
    EmailConfig,
    attachment_size_ok,
    build_backup_zip,
    build_field_assessment_package,
    build_hazard_import_template,
    build_hazard_pdf,
    build_hazard_register_workbook,
    build_orl_assessment_workbook,
    build_orl_pdf,
    build_requirements_import_template,
    reset_assessment_state,
    send_email_backup,
    template_catalogue_csv,
    valid_email,
)

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture()
def demo_frames() -> tuple[pd.DataFrame, pd.DataFrame]:
    return (
        pd.read_csv(ROOT / "sample_data" / "requirements_sample.csv"),
        pd.read_csv(ROOT / "sample_data" / "hazards_sample.csv"),
    )


def test_introduction_and_disclaimer_configuration() -> None:
    assert APP_TITLE == "MOBRA — Mobile Operational Biosecurity Readiness Assessment"
    assert len(HOW_TO_USE_STEPS) == 6
    assert "does not constitute" in FULL_DISCLAIMER
    assert AUTHOR_EMAIL == "modiabat@gmail.com"


def test_author_email_can_be_overridden_or_disabled(monkeypatch) -> None:
    monkeypatch.setenv("MOBRA_AUTHOR_EMAIL", "feedback@example.org")
    assert configured_author_email() == "feedback@example.org"
    monkeypatch.setenv("MOBRA_AUTHOR_EMAIL", "")
    assert configured_author_email() == ""


def test_orl_workbook_has_all_sixty_requirements_and_stable_columns(demo_frames) -> None:
    requirements, _ = demo_frames
    workbook = load_workbook(io.BytesIO(build_orl_assessment_workbook(requirements)), read_only=True)
    assert workbook.sheetnames == ["Instructions", "ORL_Assessment"]
    sheet = workbook["ORL_Assessment"]
    assert sheet.max_row == 61
    assert sheet.cell(2, 1).value == "R001"
    assert sheet.cell(61, 1).value == "R060"
    assert sheet.cell(1, 1).value == "Requirement ID"


def test_requirements_import_template_has_exact_supported_columns() -> None:
    workbook = load_workbook(io.BytesIO(build_requirements_import_template()), read_only=True)
    assert workbook.sheetnames == ["Instructions", "Requirements_Import"]
    assert [cell.value for cell in workbook["Requirements_Import"][1]] == [
        "requirement_id",
        "domain",
        "requirement",
        "objective_evidence",
        "observed_score",
        "maximum_score",
        "critical_control",
        "compliance_status",
        "corrective_action",
        "responsible_person",
        "due_date",
    ]


def test_orl_pdf_is_generated_with_boundary_requirement_ids(demo_frames) -> None:
    requirements, _ = demo_frames
    pdf = build_orl_pdf(requirements)
    assert pdf.startswith(b"%PDF")
    assert b"R001" in pdf
    assert b"R060" in pdf
    assert b"Disclaimer" in pdf
    assert b"Page 1 of" in pdf
    assert pdf.count(b"Requirement ID | Domain") >= 2


def test_hazard_templates_have_required_fields_and_pdf(demo_frames) -> None:
    _, hazards = demo_frames
    workbook = load_workbook(io.BytesIO(build_hazard_register_workbook(hazards)), read_only=True)
    assert workbook.sheetnames == ["Instructions", "Hazard_Register"]
    headers = [cell.value for cell in workbook["Hazard_Register"][1]]
    assert {"Hazard ID", "Likelihood", "Consequence", "Residual Likelihood", "Residual Consequence"}.issubset(headers)
    template = load_workbook(io.BytesIO(build_hazard_import_template()), read_only=True)
    assert "hazard_id" in [cell.value for cell in template["Hazard_Import"][1]]
    assert build_hazard_pdf(hazards).startswith(b"%PDF")
    assert b"Page 1 of" in build_hazard_pdf(hazards)


def test_combined_package_has_stable_protected_sheet_names(demo_frames) -> None:
    requirements, hazards = demo_frames
    workbook = load_workbook(io.BytesIO(build_field_assessment_package(requirements, hazards)), read_only=True)
    assert workbook.sheetnames == [
        "Instructions",
        "Assessment_Metadata",
        "ORL_Assessment",
        "Hazard_Register",
        "Corrective_Action_Plan",
        "Scoring_Guidance",
        "Disclaimer",
    ]


def test_backup_zip_contains_derived_outputs_and_disclaimer() -> None:
    content = build_backup_zip({"MOBRA_Summary.json": b"{}"})
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert set(archive.namelist()) == {"README.txt", "DISCLAIMER.txt", "MOBRA_Summary.json"}
        assert b"Original uploaded source files" in archive.read("README.txt")
        assert b"experimental research" in archive.read("DISCLAIMER.txt")


def test_email_validation_and_disabled_state() -> None:
    assert valid_email("user@example.org")
    assert not valid_email("not-an-email")
    assert attachment_size_ok({"a": b"1"})
    assert not attachment_size_ok({"a": b"x" * (MAX_EMAIL_ATTACHMENT_BYTES + 1)})
    with pytest.raises(EmailBackupError, match="disabled"):
        send_email_backup(
            EmailConfig(),
            recipient="user@example.org",
            subject="MOBRA Application Inquiry",
            assessment_name="Demo",
            attachments={"summary.json": b"{}"},
            consent=True,
            authorized=True,
            no_sensitive_data=True,
        )


def test_email_requires_consent_and_rejects_invalid_recipient() -> None:
    config = EmailConfig(host="smtp.example.org", username="u", password="p", sender="from@example.org", enabled=True)
    with pytest.raises(EmailBackupError, match="valid recipient"):
        send_email_backup(
            config,
            recipient="invalid",
            subject="MOBRA Application Inquiry",
            assessment_name="Demo",
            attachments={},
            consent=True,
            authorized=True,
            no_sensitive_data=True,
        )
    with pytest.raises(EmailBackupError, match="Confirm consent"):
        send_email_backup(
            config,
            recipient="user@example.org",
            subject="MOBRA Application Inquiry",
            assessment_name="Demo",
            attachments={},
            consent=False,
            authorized=True,
            no_sensitive_data=True,
        )


def test_reset_clears_assessment_session_state() -> None:
    state = {"hazards": "uploaded", "backup_recipient": "user@example.org", "filters": ["High"]}
    reset_assessment_state(state)
    assert state == {"_mobra_reset_message": True}


def test_manuscript_is_not_fabricated_when_missing() -> None:
    assert (ROOT / "docs" / "MOBRA_Manuscript.pdf").exists()


def test_template_catalogue_has_stable_download_contract() -> None:
    catalogue = template_catalogue_csv().decode("utf-8-sig")
    assert "MOBRA_Field_Assessment_Package.xlsx" in catalogue
    assert "Ready for printing" in catalogue
    assert "Re-upload compatible" in catalogue
