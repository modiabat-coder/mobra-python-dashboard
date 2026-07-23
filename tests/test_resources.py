"""Normative-resource manifest, attribution, and catalogue tests."""

from __future__ import annotations

import io
import json
from pathlib import Path

from openpyxl import load_workbook

from mobra.resources import (
    catalogue_csv_bytes,
    catalogue_xlsx_bytes,
    load_normative_resources,
    load_supporting_literature,
    resource_catalogue_frame,
    validate_resource_manifest,
)

ROOT = Path(__file__).resolve().parents[1]


def test_normative_manifest_validates_and_has_required_resources() -> None:
    resources = load_normative_resources()
    assert len(resources) >= 9
    assert validate_resource_manifest(resources) == []
    assert {item["resource_id"] for item in resources} >= {
        "WHO-01",
        "WHO-02",
        "WHO-03",
        "WHO-04",
        "WHO-05",
        "ISO-01",
        "ISO-02",
        "ISO-03",
        "BMBL-01",
    }


def test_every_resource_has_official_page_and_authorized_download_rules() -> None:
    for resource in load_normative_resources():
        assert resource["official_page_url"].startswith("https://")
        if resource["official_download_url"]:
            assert resource["official_download_url"].startswith("https://")
        if resource["resource_id"].startswith("ISO-"):
            assert resource["redistribution_status"] == "Do not redistribute"
            assert not resource.get("local_file")


def test_duplicate_resource_ids_are_rejected() -> None:
    resources = load_normative_resources()
    duplicated = [resources[0], dict(resources[0])]
    assert any("Duplicate resource_id" in error for error in validate_resource_manifest(duplicated))


def test_catalogue_exports_preserve_citations_and_metadata() -> None:
    resources = load_normative_resources()
    frame = resource_catalogue_frame(resources)
    csv = catalogue_csv_bytes(resources).decode("utf-8-sig")
    assert "citation" in csv
    assert "WHO-05" in csv
    workbook = load_workbook(io.BytesIO(catalogue_xlsx_bytes(resources)), read_only=True)
    assert workbook.sheetnames == ["Normative_Resources"]
    assert workbook["Normative_Resources"].max_row == len(resources) + 1
    assert len(load_supporting_literature()) == 5
    assert set(frame["resource_type"]) >= {"Normative guidance", "International standard", "Advisory best practice"}


def test_missing_external_links_do_not_crash(tmp_path: Path) -> None:
    path = tmp_path / "resources.json"
    payload = {"resources": [dict(load_normative_resources()[0], official_download_url="")]}
    path.write_text(json.dumps(payload), encoding="utf-8")
    assert load_normative_resources(path)[0]["official_download_url"] == ""


def test_supporting_literature_is_not_normative() -> None:
    for item in load_supporting_literature():
        assert item["evidence_role"]
        assert item["evidence_role"] != "Normative guidance"
        assert item["official_or_publisher_url"].startswith("https://")
