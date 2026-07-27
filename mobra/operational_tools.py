"""Operational templates, printable forms, ZIP backups, and optional email delivery.

The functions in this module are intentionally independent from MOBRA scoring. They
only read the current schemas and package user-selected outputs; they never alter
the demonstration data or deployment rules.
"""

from __future__ import annotations

import hashlib
import io
import os
import re
import smtplib
import zipfile
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import UTC, datetime
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import APP_TITLE, APP_VERSION, FULL_DISCLAIMER
from .security import (
    safe_archive_name,
    spreadsheet_safe_frame,
)

EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
MAX_EMAIL_ATTACHMENT_BYTES = 20 * 1024 * 1024
SAMPLE_DIR = Path(__file__).resolve().parent.parent / "sample_data"

ORL_TEMPLATE_COLUMNS = [
    "Requirement ID",
    "Domain",
    "Requirement",
    "Observed Score",
    "Maximum Score",
    "Applicable",
    "Critical Control",
    "Objective Evidence",
    "Finding or Gap",
    "Corrective Action",
    "Responsible Person",
    "Target Date",
    "Assessor Notes",
]
HAZARD_TEMPLATE_COLUMNS = [
    "Hazard ID",
    "Hazard Description",
    "Domain",
    "Activity",
    "Cause",
    "Existing Controls",
    "Likelihood",
    "Consequence",
    "Residual Likelihood",
    "Residual Consequence",
    "Corrective Action",
    "Responsible Person",
    "Status",
    "Target Date",
]
SCORING_GUIDANCE = (
    "0 = absent or not implemented; 1 = minimal implementation; 2 = limited implementation; "
    "3 = partial implementation; 4 = substantially implemented; 5 = fully implemented with adequate evidence. "
    "This scale is provisional and has not completed scientific validation."
)

TEMPLATE_CATALOGUE = [
    {
        "filename": "MOBRA_Printable_ORL_Assessment_Form.xlsx",
        "format": "XLSX",
        "status": "Ready for digital entry",
        "reupload": "Re-upload compatible",
    },
    {
        "filename": "MOBRA_Printable_ORL_Assessment_Form.pdf",
        "format": "PDF",
        "status": "Ready for printing",
        "reupload": "Manual re-entry required",
    },
    {
        "filename": "MOBRA_Requirements_Import_Template.xlsx",
        "format": "XLSX",
        "status": "Ready for digital entry",
        "reupload": "Re-upload compatible",
    },
    {
        "filename": "MOBRA_Printable_Hazard_Register.xlsx",
        "format": "XLSX",
        "status": "Ready for digital entry",
        "reupload": "Manual re-entry required",
    },
    {
        "filename": "MOBRA_Printable_Hazard_Register.pdf",
        "format": "PDF",
        "status": "Ready for printing",
        "reupload": "Manual re-entry required",
    },
    {
        "filename": "MOBRA_Hazard_Import_Template.xlsx",
        "format": "XLSX",
        "status": "Ready for digital entry",
        "reupload": "Re-upload compatible",
    },
    {
        "filename": "MOBRA_Field_Assessment_Package.xlsx",
        "format": "XLSX",
        "status": "Ready for digital entry",
        "reupload": "Re-upload compatible",
    },
]


def _sample_requirements() -> pd.DataFrame:
    return pd.read_csv(SAMPLE_DIR / "requirements_sample.csv")


def _sample_hazards() -> pd.DataFrame:
    return pd.read_csv(SAMPLE_DIR / "hazards_sample.csv")


def _sheet_from_frame(workbook: Workbook, name: str, frame: pd.DataFrame, *, freeze: str = "A2") -> None:
    frame = spreadsheet_safe_frame(frame)
    sheet = workbook.create_sheet(name)
    for column_index, column in enumerate(frame.columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=str(column))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B3954")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in frame.itertuples(index=False, name=None):
        sheet.append(list(row))
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, column in enumerate(frame.columns, start=1):
        width = min(
            42, max(14, len(str(column)) + 2, *(len(str(value)) for value in frame.iloc[:25, column_index - 1]))
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _instruction_sheet(workbook: Workbook, title: str, lines: list[str]) -> None:
    sheet = workbook.create_sheet("Instructions", 0)
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B3954")
    for index, line in enumerate([f"Application: {APP_TITLE}", *lines], start=3):
        sheet.cell(row=index, column=1, value=line)
        sheet.cell(row=index, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 120


def _workbook_bytes(sheets: list[tuple[str, pd.DataFrame]], title: str, instructions: list[str]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _instruction_sheet(workbook, title, instructions)
    for name, frame in sheets:
        _sheet_from_frame(workbook, name, frame)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _orl_frame(requirements: pd.DataFrame | None) -> pd.DataFrame:
    data = requirements.copy() if requirements is not None else _sample_requirements()
    result = pd.DataFrame(
        {
            "Requirement ID": data.get("requirement_id", pd.Series(dtype=str)),
            "Domain": data.get("domain", pd.Series(dtype=str)),
            "Requirement": data.get("requirement", pd.Series(dtype=str)),
            "Observed Score": data.get("observed_score", pd.Series(dtype=float)),
            "Maximum Score": data.get("maximum_score", pd.Series(dtype=float)),
            "Applicable": data.get("applicable", True),
            "Critical Control": data.get("critical_control", pd.Series(dtype=bool)),
            "Objective Evidence": data.get("evidence", data.get("objective_evidence", "")),
            "Finding or Gap": "",
            "Corrective Action": data.get("corrective_action", ""),
            "Responsible Person": data.get("responsible_person", ""),
            "Target Date": data.get("due_date", ""),
            "Assessor Notes": "",
        }
    )
    return result.reindex(columns=ORL_TEMPLATE_COLUMNS).fillna("")


def _hazard_frame(hazards: pd.DataFrame | None) -> pd.DataFrame:
    data = hazards.copy() if hazards is not None else _sample_hazards()
    result = pd.DataFrame(
        {
            "Hazard ID": data.get("hazard_id", pd.Series(dtype=str)),
            "Hazard Description": data.get("hazard", ""),
            "Domain": data.get("domain", ""),
            "Activity": data.get("activity", ""),
            "Cause": data.get("cause", ""),
            "Existing Controls": data.get("existing_controls", ""),
            "Likelihood": data.get("likelihood", ""),
            "Consequence": data.get("consequence", ""),
            "Residual Likelihood": data.get("residual_likelihood", ""),
            "Residual Consequence": data.get("residual_consequence", ""),
            "Corrective Action": data.get("corrective_action", data.get("recommended_action", "")),
            "Responsible Person": data.get("responsible_person", data.get("owner", "")),
            "Status": data.get("status", ""),
            "Target Date": data.get("due_date", ""),
        }
    )
    return result.reindex(columns=HAZARD_TEMPLATE_COLUMNS).fillna("")


def build_orl_assessment_workbook(requirements: pd.DataFrame | None = None) -> bytes:
    """Build a complete 60-row ORL assessment workbook from the active schema."""
    return _workbook_bytes(
        [("ORL_Assessment", _orl_frame(requirements))],
        "MOBRA Printable ORL Assessment Form",
        [
            "Use this blank or pre-populated form to record an assessment manually, then enter the completed values into the supported digital template.",
            SCORING_GUIDANCE,
            "Do not infer Likelihood or Consequence values automatically; qualified assessors must select and document them.",
            "Template status: Ready for printing. Digital entry is supported through the ORL import template; this form requires manual re-entry.",
            "The form supports structured assessment and does not constitute scientific, clinical, regulatory, operational, or field validation.",
            "Disclaimer: " + FULL_DISCLAIMER,
        ],
    )


def build_requirements_import_template() -> bytes:
    columns = [
        "requirement_id",
        "domain",
        "requirement",
        "objective_evidence",
        "observed_score",
        "maximum_score",
        "applicable",
        "critical_control",
        "compliance_status",
        "corrective_action",
        "responsible_person",
        "due_date",
    ]
    return _workbook_bytes(
        [("Requirements_Import", pd.DataFrame(columns=columns))],
        "MOBRA Requirements Import Template",
        [
            "Use these exact supported upload columns. Keep requirement IDs unique and use R001–R060 for the demonstration schema.",
            "Observed Score and Maximum Score must be numeric for applicable requirements. Use true/false in Applicable; not-applicable rows are excluded from readiness numerators and denominators. Dates should use YYYY-MM-DD.",
            "Template status: Ready for digital entry and re-upload compatible.",
            "Validation findings are returned in the application and do not silently change source data.",
            "Template status: Ready for printing. Handwritten entries require manual re-entry.",
            "Disclaimer: " + FULL_DISCLAIMER,
        ],
    )


def build_hazard_register_workbook(hazards: pd.DataFrame | None = None) -> bytes:
    return _workbook_bytes(
        [("Hazard_Register", _hazard_frame(hazards))],
        "MOBRA Printable Hazard Register",
        [
            "Complete the paper register with a qualified assessor. Likelihood and Consequence are not inferred automatically.",
            "Use integer Likelihood and Consequence values on the configured 1–5 scale and document evidence for the selected values.",
            "Handwritten forms require manual entry into a supported digital template later; MOBRA does not promise OCR recognition.",
            "Disclaimer: " + FULL_DISCLAIMER,
        ],
    )


def build_hazard_import_template() -> bytes:
    columns = [
        "hazard_id",
        "hazard",
        "hazard_category",
        "domain",
        "activity",
        "biological_agent",
        "cause",
        "existing_controls",
        "likelihood",
        "consequence",
        "residual_likelihood",
        "residual_consequence",
        "corrective_action",
        "responsible_person",
        "status",
        "due_date",
    ]
    return _workbook_bytes(
        [("Hazard_Import", pd.DataFrame(columns=columns))],
        "MOBRA Hazard Import Template",
        [
            "Use the exact supported upload columns. Likelihood and Consequence must be selected by qualified assessors and are not inferred automatically.",
            "Residual fields are optional; incomplete pairs are reported by validation.",
            "Template status: Ready for digital entry and re-upload compatible.",
            "Disclaimer: " + FULL_DISCLAIMER,
        ],
    )


def build_orl_pdf(requirements: pd.DataFrame | None = None) -> bytes:
    frame = _orl_frame(requirements)
    lines = [
        "MOBRA Printable ORL Assessment Form",
        "Mobile Operational Biosecurity Readiness Assessment",
        "Laboratory or mission: ________________________________    Location: ________________________________",
        "Assessment date: ______________  Assessor: __________________  Reviewers: __________________________",
        "Mission type: _________________________________________    Notes: _________________________________",
        "Scoring: " + SCORING_GUIDANCE,
        "",
        "Requirement ID | Domain | Requirement | Score (0-5) | Evidence / finding / notes",
    ]
    lines.extend(
        f"{row['Requirement ID']} | {row['Domain']} | {row['Requirement']} | ____ | __________________________________________"
        for _, row in frame.iterrows()
    )
    lines.extend(["", "Disclaimer:", *FULL_DISCLAIMER.splitlines()])
    return _simple_pdf(lines, "MOBRA ORL Assessment Form", repeat_header=lines[7])


def build_hazard_pdf(hazards: pd.DataFrame | None = None) -> bytes:
    frame = _hazard_frame(hazards)
    lines = [
        "MOBRA Printable Hazard Register",
        "Mobile Operational Biosecurity Readiness Assessment",
        "Laboratory or mission: ________________________________    Location: ________________________________",
        "Assessment date: ______________  Assessor: __________________  Reviewers: __________________________",
        "Likelihood and Consequence must be selected by qualified assessors; they are not inferred automatically.",
        "",
        "Hazard ID | Hazard | Domain | Activity | Likelihood | Consequence | Controls / corrective action",
    ]
    lines.extend(
        f"{row['Hazard ID']} | {row['Hazard Description']} | {row['Domain']} | {row['Activity']} | ____ | ____ | __________________________"
        for _, row in frame.iterrows()
    )
    lines.extend(["", "Disclaimer:", *FULL_DISCLAIMER.splitlines()])
    return _simple_pdf(lines, "MOBRA Hazard Register", repeat_header=lines[6])


def build_field_assessment_package(
    requirements: pd.DataFrame | None = None, hazards: pd.DataFrame | None = None
) -> bytes:
    metadata = pd.DataFrame(
        {
            "field": [
                "Laboratory or mission",
                "Location",
                "Assessment date",
                "Assessor",
                "Reviewers",
                "Mission type",
                "Notes",
            ],
            "value": ["", "", "", "", "", "", ""],
        }
    )
    actions = pd.DataFrame(
        columns=[
            "action_id",
            "finding_or_gap",
            "corrective_action",
            "responsible_person",
            "target_date",
            "status",
            "verification_notes",
        ]
    )
    sheets = [
        ("Assessment_Metadata", metadata),
        ("ORL_Assessment", _orl_frame(requirements)),
        ("Hazard_Register", _hazard_frame(hazards)),
        ("Corrective_Action_Plan", actions),
        ("Scoring_Guidance", pd.DataFrame({"guidance": [SCORING_GUIDANCE]})),
        ("Disclaimer", pd.DataFrame({"disclaimer": [FULL_DISCLAIMER]})),
    ]
    return _workbook_bytes(
        sheets,
        "MOBRA Field Assessment Package",
        [
            "Stable sheet names: Instructions, Assessment_Metadata, ORL_Assessment, Hazard_Register, Corrective_Action_Plan, Scoring_Guidance, Disclaimer.",
            "Complete the paper or digital assessment with qualified personnel, then upload supported digital values for validation.",
            "No scientific, clinical, operational, regulatory, or field validation is implied.",
        ],
    )


def _pdf_escape(value: str) -> str:
    ascii_value = value.encode("latin-1", "replace").decode("latin-1")
    return ascii_value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _simple_pdf(
    lines: list[str],
    title: str,
    lines_per_page: int = 42,
    *,
    repeat_header: str | None = None,
) -> bytes:
    """Create a small, readable, dependency-free PDF with repeated page content."""
    chunk_size = max(8, lines_per_page - 3)
    chunks = [lines[index : index + chunk_size] for index in range(0, len(lines), chunk_size)] or [[title]]
    pages = [
        [title, f"Page {page_number} of {len(chunks)}", *([repeat_header] if repeat_header else []), *chunk]
        for page_number, chunk in enumerate(chunks, start=1)
    ]
    objects: list[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    page_ids = [4 + 2 * index for index in range(len(pages))]
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>".encode("ascii"))
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    for page_index, page_lines in enumerate(pages):
        content_lines = ["BT", "/F1 8 Tf", "50 760 Td"]
        for line_index, line in enumerate(page_lines):
            if line_index:
                content_lines.append("0 -16 Td")
            content_lines.append(f"({_pdf_escape(str(line))}) Tj")
        content_lines.append("ET")
        stream = "\n".join(content_lines).encode("latin-1", "replace")
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 3 0 R >> >> /Contents {5 + 2 * page_index} 0 R >>".encode(
                "ascii"
            )
        )
        objects.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")
    result = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for number, obj in enumerate(objects, start=1):
        offsets.append(len(result))
        result.extend(f"{number} 0 obj\n".encode("ascii"))
        result.extend(obj)
        result.extend(b"\nendobj\n")
    xref = len(result)
    result.extend(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii"))
    for offset in offsets[1:]:
        result.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    result.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R /Info << /Title ({_pdf_escape(title)}) >> >>\nstartxref\n{xref}\n%%EOF\n".encode(
            "latin-1", "replace"
        )
    )
    return bytes(result)


@dataclass(frozen=True)
class EmailConfig:
    host: str = ""
    port: int = 587
    username: str = ""
    password: str = ""
    sender: str = ""
    use_tls: bool = True
    enabled: bool = False

    @classmethod
    def from_mapping(cls, values: Mapping[str, object] | None = None) -> EmailConfig:
        source = dict(values or {})

        def get(name: str, default: object = "") -> object:
            return source.get(name, os.getenv(name, default))

        try:
            port = int(get("MOBRA_SMTP_PORT", 587))
        except (TypeError, ValueError):
            port = 587
        return cls(
            host=str(get("MOBRA_SMTP_HOST", "") or ""),
            port=port,
            username=str(get("MOBRA_SMTP_USERNAME", "") or ""),
            password=str(get("MOBRA_SMTP_PASSWORD", "") or ""),
            sender=str(get("MOBRA_SMTP_FROM", "") or ""),
            use_tls=str(get("MOBRA_SMTP_USE_TLS", "true")).lower() not in {"0", "false", "no", "off"},
            enabled=str(get("MOBRA_EMAIL_ENABLED", "false")).lower() in {"1", "true", "yes", "on"},
        )

    @property
    def configured(self) -> bool:
        return bool(self.enabled and self.host and self.sender and self.username and self.password)


class EmailBackupError(ValueError):
    """Safe user-facing email error without exposing SMTP internals."""


def valid_email(value: str) -> bool:
    return bool(EMAIL_PATTERN.fullmatch(value.strip()))


def attachment_size_ok(attachments: Mapping[str, bytes], limit: int = MAX_EMAIL_ATTACHMENT_BYTES) -> bool:
    return sum(len(content) for content in attachments.values()) <= limit


def send_email_backup(
    config: EmailConfig,
    *,
    recipient: str,
    cc: str = "",
    subject: str,
    assessment_name: str,
    attachments: Mapping[str, bytes],
    consent: bool,
    authorized: bool,
    no_sensitive_data: bool,
) -> None:
    """Send only after explicit consent and authorization checks."""
    if not config.configured:
        raise EmailBackupError("Email backup is disabled because SMTP settings are not configured.")
    if not valid_email(recipient) or (cc and not valid_email(cc)):
        raise EmailBackupError("Enter a valid recipient and optional CC email address.")
    if not assessment_name.strip():
        raise EmailBackupError("Enter an assessment or mission name before sending.")
    if not (consent and authorized and no_sensitive_data):
        raise EmailBackupError("Confirm consent, transmission authorization, and data classification before sending.")
    if not attachment_size_ok(attachments):
        raise EmailBackupError("The selected attachments exceed the total email-size limit.")
    message = EmailMessage()
    message["From"] = config.sender
    message["To"] = recipient
    if cc:
        message["Cc"] = cc
    message["Subject"] = subject or "MOBRA Application Inquiry"
    message.set_content(
        f"{APP_TITLE} assessment backup: {assessment_name}\n\n"
        "This message was sent after explicit user consent. Verify institutional authorization and data classification.\n\n"
        "Passing software checks does not constitute scientific, clinical, operational, regulatory, institutional, or field validation."
    )
    for filename, content in attachments.items():
        message.add_attachment(content, maintype="application", subtype="octet-stream", filename=filename)
    try:
        with smtplib.SMTP(config.host, config.port, timeout=20) as server:
            if config.use_tls:
                server.starttls()
            server.login(config.username, config.password)
            server.send_message(message)
    except Exception as exc:  # pragma: no cover - integration boundary, never show raw SMTP details
        raise EmailBackupError("Email delivery failed; check SMTP configuration and recipient settings.") from exc


def build_backup_zip(files: Mapping[str, bytes]) -> bytes:
    """Package only selected derived outputs; uploaded source files are never added implicitly."""
    buffer = io.BytesIO()
    timestamp = datetime.now(UTC).replace(microsecond=0).isoformat()
    normalized_files: dict[str, bytes] = {}
    for filename, content in files.items():
        safe_name = safe_archive_name(filename)
        if safe_name in {"README.txt", "DISCLAIMER.txt"}:
            raise ValueError(f"Reserved backup filename: {safe_name}.")
        if safe_name in normalized_files:
            raise ValueError(f"Duplicate backup filename: {safe_name}.")
        normalized_files[safe_name] = content
    checksums = {
        filename: hashlib.sha256(content).hexdigest()
        for filename, content in normalized_files.items()
    }
    listing = "\n".join(
        f"- {filename} ({len(normalized_files[filename]):,} bytes, SHA-256 {checksums[filename]})"
        for filename in normalized_files
    )
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "README.txt",
            f"{APP_TITLE}\n"
            f"Application version: {APP_VERSION}\n"
            f"Created (UTC): {timestamp}\n"
            "MOBRA derived assessment backup. Original uploaded source files are excluded by default.\n"
            "Included files and integrity checksums:\n"
            f"{listing}\n\n"
            "This package contains derived outputs selected by the user; it is not an authorization to deploy.\n",
        )
        archive.writestr("DISCLAIMER.txt", FULL_DISCLAIMER)
        for filename, content in normalized_files.items():
            archive.writestr(filename, content)
    return buffer.getvalue()


def template_catalogue_csv() -> bytes:
    """Return a stable manifest for printable and digital templates."""
    return (
        spreadsheet_safe_frame(pd.DataFrame(TEMPLATE_CATALOGUE))
        .to_csv(index=False)
        .encode("utf-8-sig")
    )


def reset_assessment_state(session_state: Mapping[str, object]) -> None:
    """Clear assessment state without touching files outside application-controlled state."""
    preserve = {"_mobra_reset_message"}
    for key in list(session_state.keys()):
        if key not in preserve:
            del session_state[key]
    session_state["_mobra_reset_message"] = True
