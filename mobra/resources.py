"""Normative evidence manifest and supporting-literature catalogue helpers."""

from __future__ import annotations

import io
import json
import zipfile
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pandas as pd

from .config import PROJECT_ROOT
from .security import (
    resolve_within,
    safe_archive_name,
    spreadsheet_safe_frame,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESOURCE_REQUIRED_FIELDS = (
    "resource_id",
    "title",
    "issuing_organization",
    "edition",
    "publication_year",
    "resource_type",
    "topic",
    "official_page_url",
    "official_download_url",
    "access_type",
    "redistribution_status",
    "licence_or_copyright",
    "citation",
    "relevance_to_mobra",
    "current_status",
    "last_verified_date",
    "notes",
)
ALLOWED_ACCESS_TYPES = {
    "Free official download",
    "Official webpage",
    "Licensed purchase or institutional access",
    "Open-access scientific article",
    "Subscription or publisher access",
}
ALLOWED_REDISTRIBUTION_STATUSES = {
    "Link only",
    "Redistribution permitted with attribution",
    "Do not redistribute",
    "Licence review required",
}


def manifest_path() -> Path:
    return Path(__file__).resolve().parent.parent / "config" / "normative_resources.json"


def literature_path() -> Path:
    return Path(__file__).resolve().parent.parent / "config" / "supporting_literature.json"


def load_normative_resources(path: Path | None = None) -> list[dict[str, Any]]:
    source = path or manifest_path()
    payload = json.loads(source.read_text(encoding="utf-8"))
    resources = payload.get("resources", payload) if isinstance(payload, dict) else payload
    if not isinstance(resources, list):
        raise ValueError("Normative resource manifest must contain a list of resources.")
    errors = validate_resource_manifest(resources)
    if errors:
        raise ValueError("; ".join(errors))
    return resources


def validate_resource_manifest(resources: Iterable[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    seen: set[str] = set()
    for index, resource in enumerate(resources, start=1):
        missing = [field for field in RESOURCE_REQUIRED_FIELDS if field not in resource]
        if missing:
            errors.append(f"Resource {index} is missing fields: {', '.join(missing)}")
            continue
        resource_id = str(resource["resource_id"])
        if resource_id in seen:
            errors.append(f"Duplicate resource_id: {resource_id}")
        seen.add(resource_id)
        page_url = str(resource["official_page_url"])
        if not page_url.startswith("https://"):
            errors.append(f"{resource_id} must have an HTTPS official_page_url")
        download_url = str(resource["official_download_url"] or "")
        if download_url and not download_url.startswith("https://"):
            errors.append(f"{resource_id} must have an HTTPS official_download_url")
        if resource["access_type"] not in ALLOWED_ACCESS_TYPES:
            errors.append(f"{resource_id} has an unsupported access_type")
        if resource["redistribution_status"] not in ALLOWED_REDISTRIBUTION_STATUSES:
            errors.append(f"{resource_id} has an unsupported redistribution_status")
        if (
            str(resource["issuing_organization"]).startswith("International Organization for Standardization")
            and resource["redistribution_status"] != "Do not redistribute"
        ):
            errors.append(f"{resource_id} ISO resources must be labelled Do not redistribute")
    return errors


def load_supporting_literature(path: Path | None = None) -> list[dict[str, Any]]:
    source = path or literature_path()
    payload = json.loads(source.read_text(encoding="utf-8"))
    literature = payload.get("literature", payload) if isinstance(payload, dict) else payload
    if not isinstance(literature, list):
        raise ValueError("Supporting-literature catalogue must contain a list.")
    return literature


def resource_catalogue_frame(resources: Iterable[dict[str, Any]] | None = None) -> pd.DataFrame:
    records = list(resources or load_normative_resources())
    return pd.DataFrame(records, columns=list(RESOURCE_REQUIRED_FIELDS))


def catalogue_csv_bytes(resources: Iterable[dict[str, Any]] | None = None) -> bytes:
    return (
        spreadsheet_safe_frame(resource_catalogue_frame(resources))
        .to_csv(index=False)
        .encode("utf-8-sig")
    )


def catalogue_xlsx_bytes(resources: Iterable[dict[str, Any]] | None = None) -> bytes:
    frame = spreadsheet_safe_frame(resource_catalogue_frame(resources))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Normative_Resources"
    for column_index, column in enumerate(frame.columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B3954")
    for row in frame.itertuples(index=False, name=None):
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(frame.columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(55, max(16, len(column) + 2))
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_open_access_reference_package(resources: Iterable[dict[str, Any]] | None = None) -> bytes:
    """Bundle only explicitly supplied, permitted local files; never bundle ISO documents."""
    buffer = io.BytesIO()
    records = list(resources or load_normative_resources())
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        included = 0
        for resource in records:
            local_file = str(resource.get("local_file", ""))
            if (
                local_file
                and resource.get("redistribution_status") == "Redistribution permitted with attribution"
                and not str(resource.get("issuing_organization", "")).startswith(
                    "International Organization for Standardization"
                )
            ):
                path = resolve_within(PROJECT_ROOT, local_file)
                if path.is_file():
                    archive.write(
                        path,
                        arcname=safe_archive_name(
                            path.relative_to(PROJECT_ROOT).as_posix()
                        ),
                    )
                    included += 1
        archive.writestr(
            "README.txt",
            "MOBRA open-access reference package. Only explicitly permitted local files are included. "
            f"Files included: {included}. ISO PDFs are never bundled.\n"
            "MOBRA does not claim endorsement, certification, accreditation, or validation by any issuing organization.\n",
        )
    return buffer.getvalue()
