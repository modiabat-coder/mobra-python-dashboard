"""Auditable, user-safe input/output helpers for approved tabular files."""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd

from .validation_findings import FindingCollector, ValidationFinding

SUPPORTED_EXTENSIONS = (".csv", ".xlsx", ".xls")
DEFAULT_MAX_FILE_SIZE_BYTES = 20 * 1024 * 1024


class FileValidationError(ValueError):
    """A user-facing file error with a stable validation code."""

    def __init__(self, code: str, message: str, findings: list[ValidationFinding] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.findings = findings or []


@dataclass
class FileReadResult:
    """File data and structured findings produced before dataset validation."""

    data: pd.DataFrame = field(default_factory=pd.DataFrame)
    findings: list[ValidationFinding] = field(default_factory=list)
    filename: str = "uploaded_file"
    sheet_name: str = ""
    encoding: str = ""
    delimiter: str = ""

    @property
    def errors(self) -> list[str]:
        return [finding.message for finding in self.findings if finding.severity == "Error"]

    @property
    def warnings(self) -> list[str]:
        return [finding.message for finding in self.findings if finding.severity == "Warning"]

    @property
    def ok(self) -> bool:
        return not self.errors


def _source_bytes(source: Any) -> bytes:
    """Return bytes from a path, uploaded-file object, or bytes-like value."""
    if isinstance(source, bytes):
        return source
    if isinstance(source, bytearray):
        return bytes(source)
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if hasattr(source, "getvalue"):
        return source.getvalue()
    if hasattr(source, "read"):
        position = source.tell() if hasattr(source, "tell") else None
        data = source.read()
        if position is not None and hasattr(source, "seek"):
            source.seek(position)
        return data
    raise TypeError("Expected a file path, bytes, or an uploaded file object.")


def source_name(source: Any, fallback: str = "uploaded_file") -> str:
    """Get a useful display name for a file-like input."""
    if hasattr(source, "name"):
        return Path(str(source.name)).name
    if isinstance(source, (str, Path)):
        return Path(source).name
    return fallback


def _file_finding(
    collector: FindingCollector,
    severity: str,
    code: str,
    filename: str,
    detail: str,
    action: str,
    *,
    original_value: object = None,
    blocks_analysis: bool = False,
) -> None:
    collector.add(
        severity,
        code,
        f"File {filename!r}: {detail}",
        record_id=filename,
        original_value=original_value,
        suggested_action=action,
        blocks_analysis=blocks_analysis,
    )


def _decode_csv(data: bytes, filename: str, collector: FindingCollector) -> tuple[str, str]:
    if b"\x00" in data:
        _file_finding(
            collector,
            "Error",
            "ENCODING_FAILURE",
            filename,
            "the content contains binary null bytes and is not a readable text CSV.",
            "Export the source as a plain UTF-8 CSV file.",
            blocks_analysis=True,
        )
        return "", ""
    failures: list[str] = []
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            text = data.decode(encoding)
            if encoding not in {"utf-8-sig", "utf-8"}:
                _file_finding(
                    collector,
                    "Warning",
                    "NON_UTF8_ENCODING",
                    filename,
                    f"the CSV was decoded using {encoding} rather than UTF-8.",
                    "Re-export as UTF-8 when practical to improve portability.",
                    original_value=encoding,
                )
            return text, encoding
        except UnicodeDecodeError as exc:
            failures.append(str(exc))
    _file_finding(
        collector,
        "Error",
        "ENCODING_FAILURE",
        filename,
        "the CSV could not be decoded with UTF-8 or the approved fallback encodings.",
        "Export the file as UTF-8 CSV and upload it again.",
        original_value="; ".join(failures[:2]),
        blocks_analysis=True,
    )
    return "", ""


def _inspect_csv_structure(
    text: str,
    filename: str,
    collector: FindingCollector,
) -> str:
    physical_lines = text.splitlines()
    first_nonblank = next((index for index, line in enumerate(physical_lines) if line.strip()), None)
    if first_nonblank is None:
        _file_finding(
            collector,
            "Error",
            "EMPTY_FILE",
            filename,
            "the CSV contains no nonblank content.",
            "Upload a CSV with a header row and at least one data row.",
            blocks_analysis=True,
        )
        return ","
    if first_nonblank != 0:
        _file_finding(
            collector,
            "Warning",
            "HEADER_NOT_FIRST_ROW",
            filename,
            f"the first nonblank line is line {first_nonblank + 1}, so the header is not on the first line.",
            "Remove leading cover text or blank lines so the header is the first line.",
            original_value=first_nonblank + 1,
        )
    sample = "\n".join(physical_lines[first_nonblank : first_nonblank + 25])
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        _file_finding(
            collector,
            "Warning",
            "AMBIGUOUS_DELIMITER",
            filename,
            "a reliable delimiter could not be inferred; comma was used.",
            "Export with one consistent comma, semicolon, tab, or pipe delimiter.",
        )
    rows = list(csv.reader(io.StringIO("\n".join(physical_lines[first_nonblank:])), delimiter=delimiter))
    nonblank_rows = [
        (position, row)
        for position, row in enumerate(rows, start=first_nonblank + 1)
        if any(cell.strip() for cell in row)
    ]
    if nonblank_rows:
        expected_width = len(nonblank_rows[0][1])
        inconsistent = [(position, len(row)) for position, row in nonblank_rows[1:] if len(row) != expected_width]
        if inconsistent:
            preview = ", ".join(f"line {line} has {width}" for line, width in inconsistent[:8])
            _file_finding(
                collector,
                "Error",
                "INCONSISTENT_ROW_WIDTH",
                filename,
                f"CSV rows do not have a consistent field count; header has {expected_width} fields, {preview}.",
                "Correct quoting or delimiters so every row has the same number of fields.",
                original_value=inconsistent,
                blocks_analysis=True,
            )
    return delimiter


def _excel_formula_findings(
    data: bytes,
    filename: str,
    sheet_name: str | int,
    collector: FindingCollector,
) -> None:
    """Report formulas without cached values; formulas are never executed."""
    if not filename.lower().endswith(".xlsx"):
        return
    try:
        from openpyxl import load_workbook

        formula_book = load_workbook(io.BytesIO(data), read_only=False, data_only=False, keep_links=False)
        cached_book = load_workbook(io.BytesIO(data), read_only=False, data_only=True, keep_links=False)
        selected = formula_book.sheetnames[int(sheet_name)] if isinstance(sheet_name, int) else str(sheet_name)
        if selected not in formula_book.sheetnames or selected not in cached_book.sheetnames:
            return
        formulas = formula_book[selected]
        cached = cached_book[selected]
        missing: list[str] = []
        for row in formulas.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cached[cell.coordinate].value is None:
                    missing.append(cell.coordinate)
        if missing:
            _file_finding(
                collector,
                "Warning",
                "FORMULA_CELL_NOT_EVALUATED",
                filename,
                f"worksheet {selected!r} has {len(missing)} formula cell(s) without cached values: {', '.join(missing[:10])}.",
                "Open the workbook in a trusted spreadsheet application, recalculate, save, and upload again; MOBRA does not execute formulas.",
                original_value=missing,
            )
        formula_book.close()
        cached_book.close()
    except Exception:
        # Workbook corruption is handled by the primary reader. Formula inspection
        # must never turn an otherwise readable file into a crash or execute code.
        return


def read_data_file_with_validation(
    source: Any,
    *,
    name: str | None = None,
    sheet_name: str | int = 0,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
) -> FileReadResult:
    """Read one approved file and return data plus structured file findings."""
    filename = name or source_name(source)
    collector = FindingCollector("File")
    suffix = Path(filename).suffix.lower()
    result = FileReadResult(filename=filename, sheet_name=str(sheet_name))
    if suffix not in SUPPORTED_EXTENSIONS:
        detail = (
            "XLSM files are not approved and macro content is never executed."
            if suffix == ".xlsm"
            else f"extension {suffix or '(none)'} is unsupported."
        )
        _file_finding(
            collector,
            "Error",
            "UNSUPPORTED_FILE_TYPE",
            filename,
            detail,
            "Upload an explicitly approved CSV, XLSX, or XLS file.",
            original_value=suffix,
            blocks_analysis=True,
        )
        result.findings = collector.findings
        return result
    try:
        data = _source_bytes(source)
    except (OSError, TypeError) as exc:
        _file_finding(
            collector,
            "Error",
            "FILE_READ_FAILURE",
            filename,
            f"the file bytes could not be read ({exc}).",
            "Confirm the file is accessible and upload it again.",
            blocks_analysis=True,
        )
        result.findings = collector.findings
        return result
    if len(data) == 0:
        _file_finding(
            collector,
            "Error",
            "EMPTY_FILE",
            filename,
            "the file is empty (0 bytes).",
            "Upload a nonempty file with a header and at least one data row.",
            blocks_analysis=True,
        )
        result.findings = collector.findings
        return result
    if len(data) > max_file_size_bytes:
        _file_finding(
            collector,
            "Error",
            "FILE_TOO_LARGE",
            filename,
            f"the file size is {len(data):,} bytes, above the configured {max_file_size_bytes:,}-byte limit.",
            "Reduce the file size or ask an administrator to raise the configured limit.",
            original_value=len(data),
            blocks_analysis=True,
        )
        result.findings = collector.findings
        return result

    if suffix == ".csv":
        text, encoding = _decode_csv(data, filename, collector)
        result.encoding = encoding
        if not text:
            result.findings = collector.findings
            return result
        delimiter = _inspect_csv_structure(text, filename, collector)
        result.delimiter = delimiter
        if any(finding.severity == "Error" for finding in collector.findings):
            result.findings = collector.findings
            return result
        try:
            result.data = pd.read_csv(io.StringIO(text), sep=delimiter, engine="python")
        except (pd.errors.ParserError, ValueError) as exc:
            _file_finding(
                collector,
                "Error",
                "CSV_PARSE_FAILURE",
                filename,
                f"the CSV could not be parsed ({exc}).",
                "Correct inconsistent delimiters, quoting, or row widths and upload again.",
                blocks_analysis=True,
            )
    else:
        engine = "xlrd" if suffix == ".xls" else "openpyxl"
        try:
            workbook = pd.ExcelFile(io.BytesIO(data), engine=engine)
            sheets = list(workbook.sheet_names)
            if not sheets:
                _file_finding(
                    collector,
                    "Error",
                    "NO_READABLE_SHEETS",
                    filename,
                    "the workbook has no readable worksheets.",
                    "Add a visible worksheet containing tabular data and upload again.",
                    blocks_analysis=True,
                )
            if len(sheets) != len(set(sheets)):
                _file_finding(
                    collector,
                    "Error",
                    "DUPLICATE_SHEET_NAME",
                    filename,
                    "the workbook contains duplicate worksheet names.",
                    "Rename worksheets so every sheet name is unique.",
                    original_value=sheets,
                    blocks_analysis=True,
                )
            selected = sheets[int(sheet_name)] if isinstance(sheet_name, int) and sheets else str(sheet_name)
            if sheets and selected not in sheets:
                _file_finding(
                    collector,
                    "Error",
                    "UNKNOWN_SHEET",
                    filename,
                    f"worksheet {selected!r} does not exist; available sheets are {', '.join(sheets)}.",
                    "Select an available worksheet containing the actual data.",
                    blocks_analysis=True,
                )
            if not any(finding.severity == "Error" for finding in collector.findings):
                result.sheet_name = selected
                result.data = pd.read_excel(workbook, sheet_name=selected)
                _excel_formula_findings(data, filename, selected, collector)
        except (zipfile.BadZipFile, KeyError) as exc:
            _file_finding(
                collector,
                "Error",
                "CORRUPTED_WORKBOOK",
                filename,
                f"the workbook container is corrupted or is not a valid {suffix.upper()} file ({exc}).",
                "Open and re-save the workbook in a trusted spreadsheet application, then upload it again.",
                blocks_analysis=True,
            )
        except (OSError, ValueError, ImportError) as exc:
            lowered = str(exc).lower()
            code = (
                "PASSWORD_PROTECTED_WORKBOOK"
                if any(token in lowered for token in ("password", "encrypted", "encryption"))
                else "CORRUPTED_WORKBOOK"
            )
            detail = (
                "the workbook appears password-protected or encrypted"
                if code == "PASSWORD_PROTECTED_WORKBOOK"
                else "the workbook could not be read"
            )
            _file_finding(
                collector,
                "Error",
                code,
                filename,
                f"{detail} ({exc}).",
                (
                    "Remove workbook encryption and re-save as a valid XLSX/XLS file."
                    if code == "PASSWORD_PROTECTED_WORKBOOK"
                    else "Open and re-save the workbook, then upload it again."
                ),
                blocks_analysis=True,
            )

    if result.data.empty and not any(finding.code == "EMPTY_FILE" for finding in collector.findings):
        _file_finding(
            collector,
            "Error",
            "EMPTY_DATASET",
            filename,
            "the selected file or worksheet has no data rows.",
            "Select the worksheet containing the actual data and at least one row.",
            blocks_analysis=True,
        )
    result.findings = collector.findings
    return result


def list_excel_sheets(source: Any) -> list[str]:
    """Return worksheet names for approved XLS/XLSX input with safe errors."""
    name = source_name(source)
    suffix = Path(name).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        return []
    data = _source_bytes(source)
    if not data:
        raise FileValidationError("EMPTY_FILE", f"File {name!r} is empty.")
    engine = "xlrd" if suffix == ".xls" else "openpyxl"
    try:
        return list(pd.ExcelFile(io.BytesIO(data), engine=engine).sheet_names)
    except Exception as exc:
        lowered = str(exc).lower()
        code = (
            "PASSWORD_PROTECTED_WORKBOOK"
            if any(token in lowered for token in ("password", "encrypted"))
            else "CORRUPTED_WORKBOOK"
        )
        raise FileValidationError(code, f"File {name!r} could not be inspected: {exc}") from exc


def read_data_file(
    source: Any,
    *,
    name: str | None = None,
    sheet_name: str | int = 0,
    max_file_size_bytes: int = DEFAULT_MAX_FILE_SIZE_BYTES,
) -> pd.DataFrame:
    """Backward-compatible reader that raises one user-safe structured error."""
    result = read_data_file_with_validation(
        source,
        name=name,
        sheet_name=sheet_name,
        max_file_size_bytes=max_file_size_bytes,
    )
    if result.errors:
        first = next(finding for finding in result.findings if finding.severity == "Error")
        raise FileValidationError(first.code, first.message, result.findings)
    return result.data


def split_unified_file(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split a unified file into hazards and requirements without changing input values."""
    from .validation import normalise_columns

    normalized = normalise_columns(df)
    type_column = next((column for column in ("record_type", "record_kind", "dataset") if column in normalized), None)
    if type_column:
        values = normalized[type_column].astype(str).str.strip().str.lower()
        hazard_mask = values.str.contains("hazard|risk", regex=True, na=False)
        requirement_mask = values.str.contains("requirement|orl|control", regex=True, na=False)
        hazards = normalized.loc[hazard_mask].copy()
        requirements = normalized.loc[requirement_mask].copy()
        if hazards.empty or requirements.empty:
            raise ValueError("Unified record_type must contain at least one hazard row and one requirement row.")
        return hazards, requirements

    hazard_columns = {"hazard", "likelihood", "consequence"}
    requirement_columns = {"requirement", "observed_score", "maximum_score"}
    has_hazard = hazard_columns.issubset(normalized.columns)
    has_requirement = requirement_columns.issubset(normalized.columns)
    if has_hazard and has_requirement:
        hazard_mask = normalized["hazard"].notna() | normalized["likelihood"].notna()
        requirement_mask = normalized["requirement"].notna() | normalized["observed_score"].notna()
        hazards = normalized.loc[hazard_mask].copy()
        requirements = normalized.loc[requirement_mask].copy()
        if hazards.empty or requirements.empty:
            raise ValueError("Unified file inference found no rows for one of the two required record types.")
        return hazards, requirements
    raise ValueError(
        "Unified files need a record_type column (hazard/requirement) or both hazard and requirement field sets."
    )
